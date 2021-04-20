import os
from openpyxl import Workbook
import dbUtil as dbUtil


def read_symptoms(severity):
    if (severity == "serious"):
        file_directory = 'serious_symptoms.txt'
    elif (severity == "common"):
        file_directory = 'common_symptoms.txt'
    elif (severity == "less-common"):
        file_directory = 'less_common_symptoms.txt'

    with open(file_directory) as f:
        symptoms = f.readlines()
        # you may also want to remove whitespace characters like `\n` at the end of each line
        symptoms_list = [x.rstrip() for x in symptoms] 
        return symptoms_list

def add_symptom(symptom,severity):
    if (severity == "serious"):
        file_directory = 'serious_symptoms.txt'
    elif (severity == "common"):
        file_directory = 'common_symptoms.txt'
    elif (severity == "less-common"):
        file_directory = 'less_common_symptoms.txt'
    # Open a file with access mode 'a'
    file_object = open(file_directory, 'a')
    # Append 'hello' at the end of file
    file_object.write(symptom + "\n")
    # Close the file
    file_object.close()
    # prologUtil.assert_symptom()


def read_all_symptoms():
    serious = read_symptoms("serious")
    common = read_symptoms("common")
    less_common = read_symptoms("less-common")

    all_symptoms = serious + common + less_common
    all_symptoms[:] = [x for x in all_symptoms if x.strip()]
    return all_symptoms

def read_ulhi():
     with open('underlying_health_issues.txt') as f:
        ulhi = f.readlines()
        # you may also want to remove whitespace characters like `\n` at the end of each line
        ulhi_list = [x.rstrip() for x in ulhi] 
        print(ulhi_list)
        return ulhi_list

def diagnoses_from_db_to_excel():
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Diagnoses"
    col_names = dbUtil.get_column_names() 
    diagnoses = dbUtil.get_diagnoses()
    col_count = len(col_names)
    for i,x in enumerate(col_names):
        col_names[i] = x.replace("_", " ").title()
    sheet.append(col_names)
    
    iter = 1
    for x,diagnosis in enumerate(diagnoses):
        offset = 1
        for i in range(col_count):
            print(diagnosis.first_name)
            sheet.cell(row=offset+iter,column=1).value = diagnosis.id
            sheet.cell(row=offset+iter,column=2).value = diagnosis.first_name
            sheet.cell(row=offset+iter,column=3).value = diagnosis.last_name
            sheet.cell(row=offset+iter,column=4).value = diagnosis.email
            sheet.cell(row=offset+iter,column=5).value = diagnosis.temperature
            sheet.cell(row=offset+iter,column=6).value = diagnosis.age
            sheet.cell(row=offset+iter,column=7).value = diagnosis.symptoms
            sheet.cell(row=offset+iter,column=8).value = diagnosis.underlying_health_issues
            sheet.cell(row=offset+iter,column=9).value = diagnosis.total_ulhi
            sheet.cell(row=offset+iter,column=10).value = diagnosis.total_serious
            sheet.cell(row=offset+iter,column=11).value = diagnosis.total_common
            sheet.cell(row=offset+iter,column=12).value = diagnosis.total_less_common
            sheet.cell(row=offset+iter,column=13).value = diagnosis.systolic_val
            sheet.cell(row=offset+iter,column=14).value = diagnosis.diastolic_val
            sheet.cell(row=offset+iter,column=15).value = diagnosis.low_bp
            sheet.cell(row=offset+iter,column=16).value = diagnosis.current_fever
            sheet.cell(row=offset+iter,column=17).value = diagnosis.result

        iter = iter + 1

    workbook.save("data/diagnoses.xlsx")

def delete_symptoms(symptoms):
    serious = read_symptoms("serious")
    common = read_symptoms("common")
    less_common = read_symptoms("less-common")
    ulhi = read_ulhi()
    
    new_serious = []
    new_common = []
    new_less_common = []
    new_ulhi = []

#iterate through each symptom in the array passed to this function,
#then iterate through the symptoms in each text file; remove element 
 # if condition is met
    for symptom in symptoms: 
        for x in serious:
            if x == symptom:
                serious.remove(x)
        for x in common:
            if x == symptom:
                common.remove(x)
        for x in less_common:
            if x == symptom:
                less_common.remove(x)
        for x in ulhi:
             if x == symptom:
                ulhi.remove(x)
    
    print("NEW SERIOUS:",serious)
    print("NEW COMMON:",common)
    print("NEW LESS COMMON:",less_common)
    print("NEW ULHI:",ulhi)

    overwrite_symptoms(serious,"serious")
    overwrite_symptoms(common,"common")
    overwrite_symptoms(less_common,"less-common")
    overwrite_symptoms(ulhi,"ulhi")


def overwrite_symptoms(symptoms,severity): #delete symptoms by overwriting what is currently in txt file
    if (severity == "serious"):
            file_directory = 'serious_symptoms.txt'
    elif (severity == "common"):
            file_directory = 'common_symptoms.txt'
    elif (severity == "less-common"):
            file_directory = 'less_common_symptoms.txt'
    elif (severity == "ulhi"):
            file_directory = 'underlying_health_issues.txt'        

    with open(file_directory, "w") as f:
        for symptom in symptoms:
            f.write(symptom + "\n")
         
            
# delete_symptoms(["Fever","Headache","Dementia","Loss of Taste"])







